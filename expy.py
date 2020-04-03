import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
from urllib3.connectionpool import xrange


def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_sheet_by_name(book1.sheetnames[0])

    print(sheet.cell(0, 0).value)
    for row in xrange(0, nrows):
        for col in xrange(0, ncols):

            sheet1.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

    return book1